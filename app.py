"""舆情监控中心 - Flask Web 应用（PostgreSQL 版）"""

import json
import uuid
import hashlib
import io
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import Flask, request, jsonify, render_template, Response, send_file
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

load_dotenv()

from yuqing_analyzer import analyze_note
from db import get_conn, init_db

app = Flask(__name__)

with app.app_context():
    init_db()

# 内存暂存：上传解析后的行数据（分析完即可丢弃）
analysis_store = {}


# ── 工具函数 ──────────────────────────────────────────────

def note_key(title: str, content: str, url: str = "") -> str:
    if url:
        return hashlib.md5(url.encode()).hexdigest()
    return hashlib.md5(f"{title}\n{content}".encode()).hexdigest()


def calc_influence(row_data, col_map):
    likes    = int(row_data[col_map["点赞量"]]  or 0)
    favs     = int(row_data[col_map["收藏量"]]  or 0)
    comments = int(row_data[col_map["评论量"]]  or 0)
    shares   = int(row_data[col_map["分享量"]]  or 0)
    return likes + favs * 2 + comments * 3 + shares * 4


def influence_level(score):
    if score > 1000: return "高"
    if score >= 300: return "中"
    return "低"


def read_excel_notes(filepath):
    filepath = str(filepath)
    if filepath.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(filepath)
        ws = book.sheet_by_index(0)
        headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
        rows = [
            [ws.cell_value(r, c) for c in range(ws.ncols)]
            for r in range(1, ws.nrows)
        ]
    else:
        wb = load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [
            [cell.value for cell in ws[row_idx]]
            for row_idx in range(2, ws.max_row + 1)
        ]
    col_map = {name: idx for idx, name in enumerate(headers)}
    required = ("笔记标题", "笔记内容", "笔记话题", "点赞量", "收藏量", "评论量", "分享量")
    missing = [c for c in required if c not in col_map]
    if missing:
        raise ValueError(f"Excel 缺少必要列: {', '.join(missing)}")
    return headers, col_map, rows


# ── 路由 ──────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "请上传 Excel 文件"}), 400

    file_bytes = f.read()
    file_hash = hashlib.md5(file_bytes).hexdigest()

    # 检测重复：同一文件内容已分析过
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT file_id, filename, analyzed_at FROM batches WHERE file_hash=%s",
                (file_hash,)
            )
            existing = cur.fetchone()
    if existing:
        return jsonify({
            "already_analyzed": True,
            "file_id": existing["file_id"],
            "filename": existing["filename"],
            "analyzed_at": existing["analyzed_at"].strftime("%Y-%m-%d %H:%M") if existing["analyzed_at"] else "",
        })

    file_id = uuid.uuid4().hex[:12]
    ext = Path(f.filename).suffix.lower() or ".xlsx"
    tmp_path = Path("/tmp") / f"{file_id}{ext}"
    tmp_path.write_bytes(file_bytes)

    try:
        headers, col_map, rows = read_excel_notes(tmp_path)
    except ValueError as e:
        tmp_path.unlink(missing_ok=True)
        return jsonify({"error": str(e)}), 400
    finally:
        tmp_path.unlink(missing_ok=True)

    analysis_store[file_id] = {
        "filename": f.filename,
        "file_hash": file_hash,
        "headers": headers,
        "col_map": col_map,
        "rows": rows,
        "results": [],
        "status": "uploaded",
    }
    return jsonify({"file_id": file_id, "total": len(rows), "filename": f.filename})


@app.route("/analyze/<file_id>")
def analyze(file_id):
    store = analysis_store.get(file_id)
    if not store:
        return jsonify({"error": "文件不存在"}), 404

    col_map = store["col_map"]  # 提升到 analyze() 作用域，_build_entry 和 generate() 均可访问

    def _build_entry(i, row_data, title, content, topics, note_url, key, result):
        score = calc_influence(row_data, col_map)
        lvl   = influence_level(score)
        return {
            "index": i, "title": title, "content": content,
            "topics": topics, "note_url": note_url,
            "likes": int(row_data[col_map["点赞量"]] or 0),
            "favs":  int(row_data[col_map["收藏量"]] or 0),
            "comments": int(row_data[col_map["评论量"]] or 0),
            "shares":   int(row_data[col_map["分享量"]] or 0),
            "influence_score": score, "influence_level": lvl,
            "note_key": key, **result,
        }

    def _analyze_uncached(i, row_data, title, content, topics, note_url, key):
        """未命中缓存的笔记：调用 AI 并写入缓存。"""
        result = analyze_note(title, content, topics)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO note_cache(note_key,risk_level,risk_reason,report_category,report_text)
                    VALUES(%s,%s,%s,%s,%s)
                    ON CONFLICT(note_key) DO UPDATE SET
                      risk_level=EXCLUDED.risk_level,
                      risk_reason=EXCLUDED.risk_reason,
                      report_category=EXCLUDED.report_category,
                      report_text=EXCLUDED.report_text
                """, (key, result["risk_level"], result.get("risk_reason",""),
                      result.get("report_category",""), result.get("report_text","")))
            conn.commit()
        return _build_entry(i, row_data, title, content, topics, note_url, key, result), title, result["risk_level"]

    def generate():
        rows = store["rows"]
        total = len(rows)
        store["status"] = "analyzing"
        store["results"] = [None] * total
        done_count = 0

        # 预处理所有行，计算 key
        row_metas = []
        for i, row_data in enumerate(rows):
            title    = str(row_data[col_map["笔记标题"]] or "")
            content  = str(row_data[col_map["笔记内容"]] or "")
            topics   = str(row_data[col_map["笔记话题"]] or "")
            note_url = str(row_data[col_map["笔记链接"]] or "") if "笔记链接" in col_map else ""
            key = note_key(title, content, note_url)
            row_metas.append((i, row_data, title, content, topics, note_url, key))

        # P1-4：单次批量查缓存，避免每行独立查询
        all_keys = [m[6] for m in row_metas]
        cache_map = {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT note_key,risk_level,risk_reason,report_category,report_text "
                    "FROM note_cache WHERE note_key=ANY(%s) AND risk_level!='分析失败'",
                    (all_keys,)
                )
                for r in cur.fetchall():
                    cache_map[r["note_key"]] = dict(r)

        cached_metas   = [(m, cache_map[m[6]]) for m in row_metas if m[6] in cache_map]
        uncached_metas = [m for m in row_metas if m[6] not in cache_map]

        # 缓存命中的立即产出 SSE 事件
        for (i, row_data, title, content, topics, note_url, key), cached_result in cached_metas:
            entry = _build_entry(i, row_data, title, content, topics, note_url, key, cached_result)
            store["results"][i] = entry
            done_count += 1
            yield f"data: {json.dumps({'current':done_count,'total':total,'title':title[:50],'risk_level':cached_result['risk_level'],'cached':True}, ensure_ascii=False)}\n\n"

        # 未命中的并发调用 AI
        if uncached_metas:
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {executor.submit(_analyze_uncached, *m): m[0] for m in uncached_metas}
                for future in as_completed(futures):
                    entry, title, risk_level = future.result()
                    store["results"][entry["index"]] = entry
                    done_count += 1
                    yield f"data: {json.dumps({'current':done_count,'total':total,'title':title[:50],'risk_level':risk_level,'cached':False}, ensure_ascii=False)}\n\n"

        store["results"] = [r for r in store["results"] if r is not None]

        # 排序
        risk_order = {"高风险":0,"中风险":1,"低风险":2,"无风险":3,"分析失败":4}
        store["results"].sort(key=lambda x:(risk_order.get(x["risk_level"],5),-x["influence_score"]))
        store["status"] = "done"

        # 写入数据库
        risk_counts = {}
        for r in store["results"]:
            risk_counts[r["risk_level"]] = risk_counts.get(r["risk_level"], 0) + 1

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO batches(file_id,filename,total,risk_counts,file_hash)
                    VALUES(%s,%s,%s,%s,%s) ON CONFLICT(file_id) DO NOTHING
                """, (file_id, store["filename"], total, json.dumps(risk_counts, ensure_ascii=False), store.get("file_hash")))
                for r in store["results"]:
                    cur.execute("""
                        INSERT INTO notes(file_id,idx,title,content,topics,note_url,
                            likes,favs,comments,shares,influence_score,influence_level,
                            risk_level,risk_reason,report_category,report_text,note_key)
                        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT(file_id,idx) DO NOTHING
                    """, (file_id, r["index"], r["title"], r["content"], r["topics"],
                          r["note_url"], r["likes"], r["favs"], r["comments"], r["shares"],
                          r["influence_score"], r["influence_level"], r["risk_level"],
                          r.get("risk_reason",""), r.get("report_category",""),
                          r.get("report_text",""), r["note_key"]))
            conn.commit()

        yield f"data: {json.dumps({'done':True}, ensure_ascii=False)}\n\n"

    return Response(generate(), mimetype="text/event-stream")


@app.route("/batch/<file_id>", methods=["DELETE"])
def delete_batch(file_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM batches WHERE file_id=%s", (file_id,))
        conn.commit()
    analysis_store.pop(file_id, None)
    return jsonify({"ok": True})


@app.route("/complaints")
def complaints():
    """P1-3：直接返回所有非「待投诉」状态的笔记，供投诉跟踪抽屉使用，避免前端 N+1 查询。"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT n.file_id, n.idx, n.title, n.note_url, n.risk_level, n.report_status,
                       b.filename, b.analyzed_at
                FROM notes n
                JOIN batches b ON b.file_id = n.file_id
                WHERE n.report_status != '待投诉'
                ORDER BY b.analyzed_at DESC, n.idx
            """)
            rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "file_id":     r["file_id"],
            "idx":         r["idx"],
            "title":       r["title"],
            "note_url":    r["note_url"] or "",
            "risk_level":  r["risk_level"],
            "status":      r["report_status"],
            "filename":    r["filename"],
            "analyzed_at": r["analyzed_at"].strftime("%Y-%m-%d %H:%M") if r["analyzed_at"] else "",
        })
    return jsonify(result)


@app.route("/history")
def history():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT file_id,filename,total,risk_counts,analyzed_at FROM batches ORDER BY analyzed_at DESC LIMIT 50")
            rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "file_id": r["file_id"],
            "filename": r["filename"],
            "total": r["total"],
            "risk_counts": r["risk_counts"] or {},
            "analyzed_at": r["analyzed_at"].strftime("%Y-%m-%d %H:%M") if r["analyzed_at"] else "",
        })
    return jsonify(result)


@app.route("/results/<file_id>")
def results(file_id):
    store = analysis_store.get(file_id)
    if store and store["status"] == "done":
        return jsonify({"results": store["results"], "status": "done"})
    # 从数据库加载
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT idx,title,content,topics,note_url,likes,favs,comments,shares,
                       influence_score,influence_level,risk_level,risk_reason,
                       report_category,report_text,report_status
                FROM notes WHERE file_id=%s ORDER BY
                  CASE risk_level WHEN '高风险' THEN 0 WHEN '中风险' THEN 1
                    WHEN '低风险' THEN 2 WHEN '无风险' THEN 3 ELSE 4 END,
                  influence_score DESC
            """, (file_id,))
            rows = cur.fetchall()
    if not rows:
        return jsonify({"error": "文件不存在"}), 404
    return jsonify({"results": [dict(r) for r in rows], "status": "done"})


@app.route("/status/<file_id>/<int:note_index>", methods=["POST"])
def set_status(file_id, note_index):
    body = request.get_json()
    new_status = body.get("status", "待投诉")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE notes SET report_status=%s WHERE file_id=%s AND idx=%s",
                        (new_status, file_id, note_index))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/status/<file_id>")
def get_status(file_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT idx, report_status FROM notes WHERE file_id=%s", (file_id,))
            rows = cur.fetchall()
    return jsonify({str(r["idx"]): r["report_status"] for r in rows})


@app.route("/export/<file_id>")
def export(file_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT filename FROM batches WHERE file_id=%s", (file_id,))
            batch = cur.fetchone()
            if not batch:
                return jsonify({"error": "记录不存在"}), 404
            cur.execute("""
                SELECT idx,title,content,topics,note_url,likes,favs,comments,shares,
                       influence_score,influence_level,risk_level,risk_reason,
                       report_category,report_text,report_status
                FROM notes WHERE file_id=%s ORDER BY idx
            """, (file_id,))
            notes = [dict(r) for r in cur.fetchall()]

    wb_out = load_workbook(data_only=True) if False else __import__("openpyxl").Workbook()
    ws = wb_out.active
    headers = ["笔记标题","笔记内容","笔记话题","笔记链接","点赞量","收藏量","评论量","分享量",
               "影响力分数","影响力等级","风险等级","举报罪名","风险原因","举报文案","投诉状态"]
    ws.append(headers)

    risk_colors = {"高风险":"FF0000","中风险":"FF8C00","低风险":"000000","无风险":"000000"}
    for r in notes:
        ws.append([r["title"],r["content"],r["topics"],r["note_url"],
                   r["likes"],r["favs"],r["comments"],r["shares"],
                   r["influence_score"],r["influence_level"],r["risk_level"],
                   r["report_category"],r["risk_reason"],r["report_text"],r["report_status"]])
        color = risk_colors.get(r["risk_level"],"000000")
        for cell in ws[ws.max_row]:
            cell.font = Font(color=color)

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    orig_name = batch["filename"].rsplit(".", 1)[0]
    return send_file(buf, as_attachment=True,
                     download_name=f"{orig_name}_分析结果.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    import os
    init_db()
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5001)))
