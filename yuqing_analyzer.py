"""小红书舆情分析工具 - 检测摩天轮票务品牌侵权/假冒风险"""

import argparse
import json
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
import os
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

load_dotenv()

client = OpenAI(
    base_url=os.getenv("API_BASE_URL"),
    api_key=os.getenv("API_KEY"),
)

SYSTEM_PROMPT = """你是摩天轮票务品牌的舆情风险分析师。你的任务是分析小红书笔记内容，判断是否存在对"摩天轮票务"品牌的舆情风险，包括品牌侵权、假冒，以及损害品牌形象的负面内容。

## 重点检测行为
A. 品牌侵权类：冒充摩天轮官方、盗用品牌名义售票、假冒客服、利用品牌诈骗
B. 品牌形象损害类：公开抹黑/诋毁摩天轮、散布不实信息、夸大负面体验煽动抵制、将摩天轮等同于黄牛/诈骗等定性指控

## 风险分级标准
- 无风险：客观中性的讨论、不涉及摩天轮的内容、正面评价
- 低风险：轻微吐槽但情绪温和，影响力有限（如个人不满但未煽动他人）
- 中风险：明确负面评价且可能误导他人（如"避雷""千万别买""黄牛平台"等煽动性表述）、疑似冒用品牌、引导私下交易
- 高风险：明确假冒官方、盗用品牌诈骗、大规模煽动抵制、严重不实指控（如"诈骗平台"）

## 举报罪名（仅在有风险时从以下列表选择最合适的一个）
谩骂嘲讽、对立、虚假不实、标题党、引人不适、诱导关注、点赞、搬运抄袭他人作品、侵犯权益、色情低俗、政治敏感、涉嫌诈骗、种族歧视、违法违规、违反公德秩序、危害人身安全、涉未成年不当内容、违规营销、假货、导流到站外

## 输出格式
严格返回 JSON，不要包含其他文字：
{
    "risk_level": "无风险/低风险/中风险/高风险",
    "risk_reason": "风险原因说明",
    "report_category": "从上方举报罪名列表中选择最合适的一个（无风险时为空字符串）",
    "report_text": "举报文案（无风险时为空字符串）"
}

举报文案格式要求（仅在有风险时生成）：
- 举报类型：品牌侵权/假冒 或 不实信息/恶意诋毁（根据实际情况选择）
- 具体违规行为描述
- 涉及品牌：摩天轮票务
- 建议平台处理方式"""


def analyze_note(title: str, content: str, topics: str, retry: int = 2) -> dict:
    """调用 AI 分析单条笔记的品牌侵权风险，失败最多重试 2 次。"""
    user_msg = f"笔记标题：{title}\n笔记内容：{content}\n笔记话题：{topics}\n请以 json 格式返回分析结果。"
    for attempt in range(retry + 1):
        try:
            resp = client.chat.completions.create(
                model="kimi-k2-0905-preview",
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )
            raw = resp.choices[0].message.content.strip()
            result = json.loads(raw)
            # 校验必要字段
            for key in ("risk_level", "risk_reason", "report_category", "report_text"):
                if key not in result:
                    result[key] = ""
            return result
        except Exception as e:
            if attempt < retry:
                print(f"  [重试 {attempt + 1}/{retry}] {e}")
                time.sleep(2)
            else:
                return {
                    "risk_level": "分析失败",
                    "risk_reason": f"API调用失败: {e}",
                    "report_category": "",
                    "report_text": "",
                }


def main():
    parser = argparse.ArgumentParser(description="小红书舆情分析 - 摩天轮票务品牌侵权检测")
    parser.add_argument("input_file", help="输入 Excel 文件路径")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"文件不存在: {input_path}")
        sys.exit(1)

    # 输出文件名：原名_分析结果.xlsx
    output_path = input_path.with_name(f"{input_path.stem}_分析结果{input_path.suffix}")

    wb = load_workbook(input_path)
    ws = wb.active

    # 读取表头，建立列名到索引的映射
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(headers)}

    # 检查必要列
    engagement_cols = ("点赞量", "收藏量", "评论量", "分享量")
    for col_name in ("笔记标题", "笔记内容", "笔记话题") + engagement_cols:
        if col_name not in col_map:
            print(f"Excel 缺少必要列: {col_name}")
            sys.exit(1)

    # 追加结果列表头
    result_cols = ["影响力分数", "影响力等级", "风险等级", "举报罪名", "风险原因", "举报文案"]
    for i, name in enumerate(result_cols):
        ws.cell(row=1, column=len(headers) + 1 + i, value=name)

    total = ws.max_row - 1
    print(f"共 {total} 条笔记待分析\n")

    # 收集所有行的分析结果
    rows_with_results = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = [cell.value for cell in ws[row_idx]]
        title = str(row_data[col_map["笔记标题"]] or "")
        content = str(row_data[col_map["笔记内容"]] or "")
        topics = str(row_data[col_map["笔记话题"]] or "")

        current = row_idx - 1
        print(f"[{current}/{total}] 分析中: {title[:40]}...")

        result = analyze_note(title, content, topics)
        rows_with_results.append((row_data, result))

        level = result["risk_level"]
        if level not in ("无风险", "分析失败"):
            print(f"  >>> {level}: {result['risk_reason'][:60]}")

    # 按风险等级排序：高 > 中 > 低 > 无 > 分析失败，同级按影响力降序
    risk_order = {"高风险": 0, "中风险": 1, "低风险": 2, "无风险": 3, "分析失败": 4}

    def calc_influence(row_data):
        likes = int(row_data[col_map["点赞量"]] or 0)
        favs = int(row_data[col_map["收藏量"]] or 0)
        comments = int(row_data[col_map["评论量"]] or 0)
        shares = int(row_data[col_map["分享量"]] or 0)
        return likes + favs * 2 + comments * 3 + shares * 4

    def influence_level(score):
        if score > 1000:
            return "高"
        elif score >= 300:
            return "中"
        return "低"

    rows_with_results.sort(
        key=lambda x: (risk_order.get(x[1]["risk_level"], 5), -calc_influence(x[0]))
    )

    # 风险等级对应字体颜色
    risk_colors = {
        "高风险": Font(color="FF0000", bold=True),
        "中风险": Font(color="FF8C00"),
        "低风险": Font(color="000000"),
        "无风险": Font(color="000000"),
        "分析失败": Font(color="999999"),
    }

    # 写回排序后的数据
    for i, (row_data, result) in enumerate(rows_with_results):
        row_idx = i + 2
        font = risk_colors.get(result["risk_level"], Font(color="000000"))
        score = calc_influence(row_data)
        lvl = influence_level(score)
        # 高影响力 + 有风险 → 加粗
        if lvl == "高" and result["risk_level"] not in ("无风险", "分析失败"):
            font = Font(color=font.color.rgb if font.color else "000000", bold=True)
        # 写原始数据列
        for col_idx, val in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.font = font
        # 写分析结果列
        base = len(headers) + 1
        ws.cell(row=row_idx, column=base, value=score).font = font
        ws.cell(row=row_idx, column=base + 1, value=lvl).font = font
        ws.cell(row=row_idx, column=base + 2, value=result["risk_level"]).font = font
        ws.cell(row=row_idx, column=base + 3, value=result["report_category"]).font = font
        ws.cell(row=row_idx, column=base + 4, value=result["risk_reason"]).font = font
        ws.cell(row=row_idx, column=base + 5, value=result["report_text"]).font = font

    wb.save(output_path)
    print(f"\n分析完成，结果已保存至: {output_path}")


if __name__ == "__main__":
    main()
